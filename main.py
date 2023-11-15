import ast
import os.path
from code.argument_parser import configure_argument_parser, get_sheet
from code.constants import JSON_FILE_NAME
from code.exceptions import FileAlreadyExists, InvalidFileFormat, SheetNotFound
from code.exel import create_main_table, create_pivot_table, get_data
from code.json_parser import get_json_data

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


def main() -> None:
    args = configure_argument_parser().parse_args()
    if args.json:
        try:
            with open(args.file) as file:
                file_contents = file.read()
            if os.path.isfile(JSON_FILE_NAME) and not args.force:
                raise FileAlreadyExists(
                    f"Файл {JSON_FILE_NAME} уже существует. Переименуйте или "
                    f"удалите файл.\nПринудительно создать файл: -f/--force"
                )
            file = openpyxl.Workbook()
            sheet = file.worksheets[0]
            servers = get_json_data(ast.literal_eval(file_contents))
            print(
                'Обработано серверов:',
                sum([s.quantity for s in servers.collection.values()])
            )
            if servers.notification:
                print("\n".join(servers.notification))
            create_main_table(sheet, servers)
        except FileNotFoundError:
            print(f"Файл {args.file} не найден!")
        except UnicodeDecodeError:
            print("Файл не поддерживается")
        except SyntaxError:
            print("Синтаксическая ошибка в файле")
        except FileAlreadyExists as e:
            print(e)
        else:
            file.save(JSON_FILE_NAME)
    else:
        try:
            file = openpyxl.load_workbook(filename=args.file)
            sheet = get_sheet(file, args)
            if sheet.max_row < 2 or sheet.max_column < 9:
                raise InvalidFileFormat("Неверный формат таблицы!")
            component_set = get_data(sheet)
            component_set.repair_calculation()
        except InvalidFileException:
            print("Файл не поддерживается! Нужен файл .xlsx,.xlsm,.xltx,.xltm")
        except (SheetNotFound, InvalidFileFormat) as e:
            print(e)
        else:
            create_pivot_table(file, component_set)
            file.save(args.file)
    return


if __name__ == "__main__":
    main()
