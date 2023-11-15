from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .constants import MAIN_HEADER, MAIN_SHEET_NAME, PIVOT_HEADER, PIVOT_SHEET_NAME
from .data_structure import Component, ComponentSet, ServerSet
from .exceptions import InvalidFileFormat


def set_style(cell, header: bool = False, blank: bool = False) -> None:
    cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    cell.font = Font(name="Helvetica Neue (Headings)", size=13, color="000000")
    if header:
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF")
    elif blank:
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal="center",
        )
        cell.fill = PatternFill(fill_type="solid", start_color="DA9694")
    else:
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal="center",
        )
        cell.fill = PatternFill(fill_type="solid", start_color="DCE2F1")


def style_header(sheet: Worksheet, header: dict):
    sheet.sheet_view.zoomScale = 55
    sheet.row_dimensions[1].height = 55
    for col, title in zip(
        sheet.iter_cols(max_col=len(header), max_row=1),
        header.items(),
    ):
        sheet.column_dimensions[col[0].column_letter].width = title[1]
        col[0].value = title[0]
        set_style(col[0], header=True)


def style_line(
        sheet: Worksheet, start_line: int, length: int, empty: bool = False
):
    if empty:
        for col in sheet.iter_cols(min_row=start_line, max_row=start_line,
                                   max_col=length):
            set_style(col[0], header=True)
        sheet.merge_cells(
            start_row=start_line, end_row=start_line, start_column=1,
            end_column=length
        )
    else:
        for col in sheet.iter_cols(min_row=start_line, max_row=start_line,
                                   max_col=length):
            set_style(col[0])


def create_main_table(sheet: Worksheet, servers: ServerSet) -> None:
    style_header(sheet, MAIN_HEADER)
    line_cursor = 2
    for server in servers.collection.values():
        first_line = True
        for row, component in zip(
            sheet.iter_rows(
                max_col=len(MAIN_HEADER),
                min_row=line_cursor,
                max_row=line_cursor + len(server.config.keys()),
            ),
            server.config.values(),
        ):
            sheet.row_dimensions[row[0].row].height = 35
            if first_line:
                values = [
                    "Сервер",
                    server.model,
                    server.quantity,
                    server.sn,
                    *component.get_list(),
                    component.quantity * server.quantity,
                    "",
                ]
                first_line = False
            else:
                values = [
                    *["" for i in range(4)],
                    *component.get_list(),
                    component.quantity * server.quantity,
                    "",
                ]
            for cell, field in zip(row, values):
                cell.value = field
                if cell.column in [5, 7, 9] and not field:
                    set_style(cell, blank=True)
                else:
                    set_style(cell)
        # merge sn fields
        sheet.merge_cells(
            start_row=line_cursor,
            end_row=line_cursor + len(server.config.values()) - 1,
            start_column=4,
            end_column=4,
        )
        # blank line
        sheet.merge_cells(
            start_row=line_cursor + len(server.config),
            start_column=1,
            end_row=line_cursor + len(server.config),
            end_column=len(MAIN_HEADER),
        )
        for col in sheet.iter_cols(
            max_col=len(MAIN_HEADER), min_row=line_cursor + len(server.config)
        ):
            set_style(col[0], header=True)
        sheet.row_dimensions[line_cursor + len(server.config)].height = 35
        line_cursor += len(server.config) + 1


def create_pivot_table(file, component_set: ComponentSet):
    if PIVOT_SHEET_NAME in file.sheetnames:
        file.remove(file[PIVOT_SHEET_NAME])
    sheet = file.create_sheet(PIVOT_SHEET_NAME)
    style_header(sheet, PIVOT_HEADER)
    for row, component in zip(
        sheet.iter_rows(
            max_col=len(PIVOT_HEADER),
            min_row=2,
            max_row=len(component_set.collection.keys()) + 1,
        ),
        component_set.collection.values(),
    ):
        sheet.row_dimensions[row[0].row].height = 35
        component_str = component.get_list()
        for repair in component.for_repair.values():
            component_str.append(repair)
        component_str.append(component.comment)
        for cell in row:
            cell.value = component_str[cell.col_idx - 1]
            set_style(cell)


def get_data(sheet: Worksheet) -> ComponentSet:
    sheet.title = MAIN_SHEET_NAME
    style_header(sheet, MAIN_HEADER)

    component_set = ComponentSet()
    quantity = sheet.cell(2, 3).value
    for row in sheet.iter_rows(
        max_col=len(MAIN_HEADER), min_row=2, max_row=sheet.max_row
    ):
        if row[6].value is None:
            style_line(sheet, int(row[0].row), len(MAIN_HEADER), empty=True)
            quantity = sheet.cell(row[0].row + 1, 3).value
        else:
            style_line(sheet, int(row[0].row), len(MAIN_HEADER), empty=False)
            if sheet.cell(row[0].row, 9).value is None:
                raise InvalidFileFormat(
                    f"Не указанно количество в {row[0].row} строке!"
                )
            sheet.cell(row[0].row, 10).value = (
                    sheet.cell(row[0].row, 10).value * quantity
            )
            component = Component(
                en_name=sheet.cell(row[0].row, 5).value,
                ru_name=sheet.cell(row[0].row, 6).value,
                pn_main=sheet.cell(row[0].row, 7).value,
                pn_opt=sheet.cell(row[0].row, 8).value,
                quantity=sheet.cell(row[0].row, 10).value,
                comment=sheet.cell(row[0].row, 11).value,
            )
            component_set.add(component)
    return component_set
